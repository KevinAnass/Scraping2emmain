import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


# URL of the Jumia page to scrape
url = 'https://www.2ememain.be/l/informatique-logiciels/#q:ecran|Language:all-languages|postcode:4000'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code == 200:
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all product listings
    products = soup.find_all('li', class_='hz-Listing hz-Listing--list-item')
    result=[];

    for product in products:
        # Extract product name
        image = product.find('img').get('src')
        title  = product.find('h3', class_='hz-Listing-title').text.strip()
        description=product.find('p', class_='hz-Listing-description hz-text-paragraph').text.strip()
        price =product.find('span', class_='hz-Listing-price hz-Listing-price--desktop hz-text-price-label').text.strip()
        result.append({
            'image':image,
            'title':title,
            'description':description,
            'price':price,
        })
    
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = 'IMAGE'
    ws['B1'] = 'TITLE'
    ws['C1'] = 'DESCRIPTION'
    ws['D1'] = 'PRICE'

    for idx, entry in enumerate(result, start=2):
        ws[f'A{idx}'] = entry['image']
        ws[f'B{idx}'] = entry['title']
        ws[f'C{idx}'] = entry['description']
        ws[f'D{idx}'] = entry['price']
    
    wb.save('sample.xlsx')
    print("Excel file created and saved successfully.")



else:
    print(f"Failed to retrieve the page. Status code: {response.status_code}")
